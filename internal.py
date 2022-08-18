import json

import openpyxl
import requests
from PyQt6 import QtCore
from PyQt6.QtGui import QFont
from PyQt6.QtWidgets import QApplication, QWidget, QPushButton, QFileDialog, QListView, QProgressBar
import sys
from numba.core import event

# 窗口size
from PyQt6.QtCore import QThread, QStringListModel
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet
from requests.exceptions import MissingSchema, SSLError

ms_token = "g8vXKy2fjhjd7xrrPcCU7Wfop7isL5KAuyjofBp061Mtaxm_fA5vZ_lAlj46mvE_NnR7x-m-022QnOLdb6Em6HbYv1qm-ek2LXKHh6aTtnLpk_Ke8h7MUTkvWAUZX0cQ1JhrowY="

window_width: int = 500
window_height: int = 314

# 按钮size
button_width: int = 80
button_height: int = 20

# 边沿长度
edge_distance: int = 20

window: QWidget
crawl_button: QPushButton
import_button: QPushButton
text_box: QListView
test_button: QPushButton
progress_bar: QProgressBar
file: str
task: dict = {}
string_list_model = QStringListModel()
logList = ['[Welcome] QAQ']
string_list_model.setStringList(logList)
progress: int = 0

font: QFont = QFont("Consolas")

colors = {
    "RED": "#ff0000",
    "BLACK": "#000000"
}


# 渲染UI
def render_ui():
    # 获取窗口
    app = QApplication(sys.argv)
    global window
    window = QWindows()

    # 设置标题，尺寸和禁用拉伸


    window.resize(window_width, window_height)
    window.setWindowTitle("短视频播放量爬虫（0.0）")
    window.setFixedSize(window.width(), window.height())

    # 文本框
    global text_box
    global font
    text_box = QListView(window)
    text_box.move(edge_distance, edge_distance)
    text_box.resize(window_width - 2 * edge_distance, window_height - button_height - 3 * edge_distance)
    text_box.setModel(string_list_model)
    text_box.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

    # 爬取按钮
    global crawl_button
    crawl_button = QPushButton(window)
    crawl_button.setText("开始爬取")
    crawl_button.move(int(window_width / 3 * 2 - button_width / 2), int(window_height - edge_distance - button_height))
    crawl_button.clicked.connect(handle_crawl_button_click)

    # 导入按钮
    global import_button
    import_button = QPushButton(window)
    import_button.setText("导入表格")
    import_button.move(int(window_width / 3 * 1 - button_width / 2), int(window_height - edge_distance - button_height))
    import_button.clicked.connect(handle_import_button_click)

    # 爬取进度条
    global progress_bar
    progress_bar = QProgressBar(window)
    progress_bar.resize(window_width - 2 * edge_distance, 15)
    progress_bar.move(edge_distance, 3)
    window.show()
    sys.exit(app.exec())


def handle_crawl_button_click():
    crawl_thread = CrawlThread(window)
    crawl_thread.start()


def handle_import_button_click():
    global file
    res = QFileDialog.getOpenFileNames()[0]
    if len(res) != 0:
        file = res[0]
        log(f"[FILE_IMPORT] Path:{file}")
        try:
            global task
            task = read_xlsx()
        except InvalidFileException:
            log(f"[ERROR] File '{file}' may not a xlsx file")
        else:
            log(f"[FILE_IMPORT] Xlsx file import success")
    else:
        log(f"[FILE_IMPORT] Cancel")


def log(text: str, color: str = colors["BLACK"]):
    # todo  这里继续加字体颜色
    logList.append(text)
    string_list_model.setStringList(logList)
    text_box.setModel(string_list_model)
    text_box.verticalScrollBar().setValue(text_box.verticalScrollBar().maximum())


def request_get(url: str, allow_redirects: bool):
    headers = {
        "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36"
    }
    try:
        response = requests.get(url, allow_redirects=allow_redirects, headers=headers)
    except MissingSchema:
        log(f"[ERROR] Invalid URL '{url}', perhaps you meant 'https://{url}'")
        return None
    except SSLError:
        log(f"[ERROR] Max retries exceeded with URL '{url}'")
        pass
    else:
        if response.status_code == 200 or response.status_code == 301:
            log(f"[GET] {response.status_code}: {url}")
            return response.content
        elif response.status_code == 404:
            return 404
        else:
            log(f"[GET] {response.status_code}: {url}")
            return None


def get_url_type(url: str):
    if url.startswith("https://vm.tiktok.com") or url.startswith("https://vt.tiktok.com"):
        return "solution_1"
    elif url.startswith("https://www.tiktok.com"):
        return "solution_2"
    log(f"[ERROR] Invalid URL '{url}'")
    return ""


def solution_1(url: str):
    print(1)
    response = request_get(url=url, allow_redirects=False)
    if response is None:
        return
    soup = BeautifulSoup(response, "lxml")
    redirect_url = soup.find(name="a")["href"]
    video_id = get_tiktok_video_id(redirect_url)
    url = "https://www.tiktok.com/api/recommend/item_list/?aid=1988&app_language=es&app_name=tiktok_web&battery_info=0.44&browser_language=zh-CN&browser_name=Mozilla&browser_online=true&browser_platform=MacIntel&browser_version=5.0%20%28Macintosh%3B%20Intel%20Mac%20OS%20X%2010_15_7%29%20AppleWebKit%2F537.36%20%28KHTML%2C%20like%20Gecko%29%20Chrome%2F104.0.0.0%20Safari%2F537.36&channel=tiktok_web&cookie_enabled=true&count=30&device_id=7106463322559923714&device_platform=web_pc&focus_state=true&from_page=video&history_len=1&insertedItemID=" + video_id + "&is_fullscreen=true&is_page_visible=true&os=mac&priority_region=&referer=&region=DE&screen_height=1920&screen_width=1080&tz_name=Asia%2FShanghai&webcast_language=es&msToken=" + ms_token
    response = request_get(url=url, allow_redirects=False)
    if response is None:
        return
    rsp_json = json.loads(response)
    res = get_tiktok_data_from_api(except_id=video_id, rsp_json=rsp_json)
    if res["found"] == 1:
        log(f"[SUCCESS] Get video:{video_id}; share: {res['share']}; play: {res['play']}; like: {res['digg']}; comment: {res['comment']}")
    elif res["found"] == 0:
        log("[FAILED] This video may lose efficacy")
    else:
        log("[FAILED] Api response may issue")
    return res


def solution_2(url: str):
    print(2)
    video_id = get_tiktok_video_id(url)
    url_ = "https://www.tiktok.com/api/recommend/item_list/?aid=1988&app_language=es&app_name=tiktok_web&battery_info=0.44&browser_language=zh-CN&browser_name=Mozilla&browser_online=true&browser_platform=MacIntel&browser_version=5.0%20%28Macintosh%3B%20Intel%20Mac%20OS%20X%2010_15_7%29%20AppleWebKit%2F537.36%20%28KHTML%2C%20like%20Gecko%29%20Chrome%2F104.0.0.0%20Safari%2F537.36&channel=tiktok_web&cookie_enabled=true&count=30&device_id=7106463322559923714&device_platform=web_pc&focus_state=true&from_page=video&history_len=1&insertedItemID=" + video_id + "&is_fullscreen=true&is_page_visible=true&os=mac&priority_region=&referer=&region=DE&screen_height=1920&screen_width=1080&tz_name=Asia%2FShanghai&webcast_language=es&msToken=" + ms_token
    response = request_get(url=url_, allow_redirects=False)
    print(url_)
    print(response)
    if response is None:
        return
    rsp_json = json.loads(response)
    res = get_tiktok_data_from_api(except_id=video_id, rsp_json=rsp_json)
    if res["found"] == 1:
        log(f"[SUCCESS] Get video:{video_id}; share: {res['share']}; play: {res['play']}; like: {res['digg']}; comment: {res['comment']}")
    elif res["found"] == 0:
        log("[FAILED] This video may lose efficacy")
    else:
        log("[FAILED] Api response may issue")
    return res


def get_tiktok_video_id(url: str) -> str:
    res_begin_index: int = url.rfind("/")
    res_end_index: int = url.find("?")
    res = url[res_begin_index + 1:res_end_index]
    for letter in res:
        if letter.isalpha():
            return ""
    return res


def read_xlsx() -> dict:
    wb: Workbook = openpyxl.open(filename=file)
    ws: Worksheet = wb.worksheets[0]
    task_list = []
    row: int = 2
    while ws.cell(row, 1).value is not None:
        task_list.append(ws.cell(row, 1).value)
        row = row + 1
    return {
        'task_count': row,
        'task_list': task_list
    }


class CrawlThread(QThread):
    def run(self):
        global task
        if task == {}:
            log("[ERROR] None file imported")
        else:
            log("[BEGIN] Crawl start")
            wb = openpyxl.Workbook()
            write_res_header(wb=wb)
            global progress
            for url in task.get("task_list"):
                url_type = get_url_type(url=url)
                if url_type == "solution_1":
                    res = solution_1(url)
                elif url_type == "solution_2":
                    res = solution_2(url)
                else:
                    res ={
                        "found": -2,
                        "video_id": "",
                        "share": 0,
                        "play": 0,
                        "digg": 0,
                        "comment": 0,
                    }
                print(res)
                write_res_line(wb=wb,
                               row=progress + 2,
                               url=url,
                               video_id=res["video_id"],
                               comment=res["comment"],
                               like=res["digg"],
                               play=res["play"],
                               share=res["share"],
                               status=res["found"]
                               )
                progress_bar.setValue(int(progress / (len(task["task_list"]) - 1) * 100))
                progress = progress + 1
            log("[COMPLETE] Crawl complete")


def get_tiktok_data_from_api(except_id: str, rsp_json: dict) -> dict:
    try:
        itemList = rsp_json["itemList"]
        for item in itemList:
            if item["id"] == except_id:
                share = item["stats"]["shareCount"]
                play = item["stats"]["playCount"]
                digg = item["stats"]["diggCount"]
                comment = item["stats"]["commentCount"]
                return {
                    "found": 1,
                    "video_id": except_id,
                    "share": share,
                    "play": play,
                    "digg": digg,
                    "comment": comment
                }
        return {
            "found": 0,
            "video_id": except_id,
            "share": 0,
            "play": 0,
            "digg": 0,
            "comment": 0,
        }
    except KeyError:
        return {
            "found": -1,
            "video_id": except_id,
            "share": 0,
            "play": 0,
            "digg": 0,
            "comment": 0,
        }


def write_res_line(wb: Workbook, row: int, url: str, video_id: str, status: int, like: int, comment: int, share: int, play: int):
    status_dict = {
        -1: "联系徐硕改代码",
        1: "Success",
        0: "Video lose efficacy",
        -2: "Bad URL or this website has no solution"
    }
    wb.worksheets[0].cell(row=row, column=1, value=url)
    wb.worksheets[0].cell(row=row, column=2, value=video_id)
    wb.worksheets[0].cell(row=row, column=3, value=status_dict[status])
    wb.worksheets[0].cell(row=row, column=4, value=like)
    wb.worksheets[0].cell(row=row, column=5, value=comment)
    wb.worksheets[0].cell(row=row, column=6, value=share)
    wb.worksheets[0].cell(row=row, column=7, value=play)
    wb.save(filename="Users/rockey220224/Desktop/output.xlsx")
    log(f"[SAVE] Save id = {video_id} success")


def write_res_header(wb: Workbook):
    wb.worksheets[0].cell(row=1, column=1, value="URL")
    wb.worksheets[0].cell(row=1, column=2, value="VideoID")
    wb.worksheets[0].cell(row=1, column=3, value="Status")
    wb.worksheets[0].cell(row=1, column=4, value="LikesCount")
    wb.worksheets[0].cell(row=1, column=5, value="CommentsCount")
    wb.worksheets[0].cell(row=1, column=6, value="SharesCount")
    wb.worksheets[0].cell(row=1, column=7, value="PlaysCount")
    wb.save(filename="Users/rockey220224/Desktop/output.xlsx")
    log(f"[SAVE] Save header success")


class QWindows(QWidget):
    def enterEvent(self, a0: event) -> None:
        global text_box
        global font
        text_box = QListView(window)
        text_box.move(edge_distance, edge_distance)
        text_box.resize(window_width - 2 * edge_distance, window_height - button_height - 3 * edge_distance)
        text_box.setModel(string_list_model)
        text_box.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        global progress_bar
        progress_bar.hide()
        progress_bar = QProgressBar(window)
        progress_bar.resize(window_width - 2 * edge_distance, 15)
        progress_bar.move(edge_distance, 3)
        global progress
        if task != {}:
            progress_bar.setValue(int(progress / (len(task["task_list"]) - 1) * 100))
            print(int(progress / (len(task["task_list"]) - 1) * 100))
        progress_bar.show()
