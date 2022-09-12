import time

import openpyxl
from openpyxl import Workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from enums import XlsxWorkerStatus, XlsxReadStatus, XlsxWriteStatus, VideoResponseStatus


class ReadResult:
    status: XlsxReadStatus
    url: str

    def __init__(self, status: XlsxReadStatus, url: str):
        self.status = status
        self.url = url


class XlsxWorker:
    status: XlsxWorkerStatus
    cur_line: int
    wb: Workbook
    ws: Worksheet
    output_path: str
    column_count: int = 0

    def __init__(self):
        pass

    def read_url(self) -> ReadResult:
        if self.status != XlsxWorkerStatus.Reader:
            return ReadResult(XlsxReadStatus.PermissionDenied, "")
        value: str = self.ws.cell(self.cur_line, 1).value
        if value is None:
            return ReadResult(XlsxReadStatus.NoMoreData, "")
        else:
            self.cur_line += 1
            return ReadResult(XlsxReadStatus.Success, value)

    def writer_line(self, value: dict) -> XlsxWriteStatus:
        if self.status != XlsxWorkerStatus.Writer:
            return XlsxWriteStatus.PermissionDenied
        for key in value.keys():
            flag: bool = False
            for i in range(1, self.column_count + 1):
                if self.ws.cell(1, i).value == key:
                    if type(value[key]) is VideoResponseStatus:
                        value[key] = str(VideoResponseStatus)
                    self.ws.cell(row=self.cur_line, column=i, value=value[key])
                    flag = True
                    break
            if not flag:
                return XlsxWriteStatus.NoSuchField
        self.cur_line += 1
        self.wb.save(self.output_path)
        return XlsxWriteStatus.Success


def init_reader(path: str) -> XlsxWorker | None:
    instance = XlsxWorker()
    instance.status = XlsxWorkerStatus.Reader
    try:
        instance.wb = openpyxl.open(filename=path)
        instance.ws = instance.wb.worksheets[0]
    except InvalidFileException:
        return None
    instance.cur_line = 2
    return instance


def init_writer(path: str, fields: tuple) -> XlsxWorker:
    instance = XlsxWorker()
    instance.wb = openpyxl.Workbook()
    instance.output_path = path
    instance.status = XlsxWorkerStatus.Writer
    instance.ws = instance.wb.worksheets[0]
    column = 1
    instance.column_count = len(fields)
    for field in fields:
        instance.ws.cell(row=1, column=column, value=field)
        column += 1
    instance.wb.save(path)
    instance.cur_line = 2
    return instance
