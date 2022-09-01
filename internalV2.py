import json
import sys
import time
from threading import Thread
from enum import Enum

import requests
from PyQt6.QtCore import QObject, pyqtSignal
from PyQt6.QtWidgets import QApplication, QPushButton, QProgressBar, QListWidget, QWidget, QFileDialog, QLineEdit, QLabel
from PyQt6 import QtCore
from bs4 import BeautifulSoup
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
import openpyxl
from PyQt6.QtGui import QIntValidator
import datetime

from requests.exceptions import MissingSchema, SSLError


class ResultStatus:
    Success = "success"
    MayNetworkError = "mayNetworkError"


class UpdateUISignals(QObject):
    log_signal = pyqtSignal(str, str)
    progress_bar_update_signal = pyqtSignal()


class AdjustConfigSignals(QObject):
    adjust_output_path_signal = pyqtSignal(str)
    adjust_begin_line_signal = pyqtSignal(int)


class Feishu(QObject):
    tenant_access_token: str = ''
    robot_secret_data = '{"app_id": "cli_a24d49eae1bad013","app_secret": "6een1lFRrlxMaFJxTIRfUetTqHRklgiZ"}'
    robot_url = 'https://open.feishu.cn/open-apis/bot/v2/hook/8260d294-6983-419d-b071-fc462d36ea70'

    update_ui_signals: UpdateUISignals

    at_user_open_id = ''
    when_complete_at: bool = False

    headers = {
        "Content-Type": "application/json; charset=utf-8"
    }

    def __init__(self, update_ui_signals: UpdateUISignals):
        super().__init__()
        self.update_ui_signals = update_ui_signals

    def __get_tenant_access_token(self) -> str:
        if self.tenant_access_token != '':
            url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
            response = requests.post(url=url, data=self.robot_secret_data, headers=headers)
            response_json = json.loads(response.content)
            if response_json['code'] != 0:
                self.update_ui_signals.log_signal.emit("FEISHU", "Get feishu access token error.")
                raise Exception(f"[{response_json['code']}]Get feishu tenant access token failed!\n{response_json['msg']}")
            self.tenant_access_token = response_json["tenant_access_token"]
            self.update_ui_signals.log_signal.emit("FEISHU", "Get feishu access token success.")
        return self.tenant_access_token

    def send_lark(self, text: str, at: bool = False):
        for i in range(0, 3):
            if not at or self.at_user_open_id == "":
                if self.__send_lark(text=text, at=""):
                    break
            else:
                if self.__send_lark(text=text, at=self.at_user_open_id):
                    break

    def __send_lark(self, text: str, at: str = "") -> bool:
        if at != "":
            obj = {"msg_type": "text", "content": " {\"text\":\"<at user_id=\\\"" + at + "\\\"></at> " + text + "\"}"}
        else:
            obj = {"msg_type": "text", "content": {"text": text}}
        try:
            requests.post(self.robot_url, json=obj, timeout=60)
        except Exception as e:
            print(e)
            return False
        return True

    def get_user_open_id_by_email(self, email: str) -> str:
        headers = {
            "Content-Type": "application/json; charset=utf-8",
            "Authorization": "Bearer " + self.__get_tenant_access_token()
        }
        url = "https://open.feishu.cn/open-apis/contact/v3/users/batch_get_id"
        data = '{"emails": ["' + email + '"]}'
        response = requests.post(url=url, data=data, headers=headers)
        rsp = json.loads(response.content)
        if rsp["data"]["user_list"][0]["user_id"] is None:
            self.update_ui_signals.log_signal.emit("FEISHU", f"Set when complete at {email} failed, please check email spell.")
            self.at_user_open_id = ""
            return ""
        self.update_ui_signals.log_signal.emit("FEISHU", f"Set when complete at {email} success")
        self.at_user_open_id = rsp["data"]["user_list"][0]["user_id"]
        return self.at_user_open_id


class ConfigWindow(QWidget):
    config_window_width = 400
    config_window_height = 250
    button_width: int = 80
    button_height: int = 20
    line_edit_height: int = 20
    line_edit_position_x = 80
    config_window_edge_distance = 10
    output_line_edit_width = 300

    begin_line: int
    output_path: str

    windows: QWidget
    save_button: QPushButton
    cancel_button: QPushButton
    begin_line_edit_text: QLineEdit
    output_path_edit_text: QLineEdit
    begin_line_label: QLabel
    output_path_label: QLabel

    adjust_config_signals: AdjustConfigSignals
    update_ui_signals: UpdateUISignals

    def __init__(self, begin_line: int, output_path: str, adjust_config_signals: AdjustConfigSignals, update_ui_signals: UpdateUISignals):
        super().__init__()

        self.adjust_config_signals = adjust_config_signals
        self.update_ui_signals = update_ui_signals

        self.setWindowTitle("爬虫配置")
        self.resize(self.config_window_width, self.config_window_height)
        self.output_path = output_path
        self.begin_line = begin_line

        self.save_button = QPushButton(self)
        self.save_button.move(int(self.config_window_width / 3) - int(self.button_width / 2), self.config_window_height - self.config_window_edge_distance - self.button_height)
        self.save_button.setText("保存")
        self.save_button.clicked.connect(self.handle_save_button_click)

        self.cancel_button = QPushButton(self)
        self.cancel_button.move(int(self.config_window_width / 3 * 2) - int(self.button_width / 2), self.config_window_height - self.config_window_edge_distance - self.button_height)
        self.cancel_button.clicked.connect(self.handle_cancel_button_click)
        self.cancel_button.setText("取消")

        self.begin_line_edit_text = QLineEdit(self)
        self.begin_line_edit_text.move(self.line_edit_position_x, self.config_window_edge_distance)
        self.begin_line_edit_text.setText(str(self.begin_line))
        self.begin_line_edit_text.setValidator(QIntValidator())

        self.output_path_edit_text = QLineEdit(self)
        self.output_path_edit_text.move(self.line_edit_position_x, self.config_window_edge_distance + self.config_window_edge_distance + self.line_edit_height)
        self.output_path_edit_text.setText(self.output_path)

        self.begin_line_label = QLabel(self)
        self.begin_line_label.setText("起点行数")
        self.begin_line_label.move(self.config_window_edge_distance, self.config_window_edge_distance)

        self.output_path_label = QLabel(self)
        self.output_path_label.setText("输出路径")
        self.output_path_label.move(self.config_window_edge_distance, self.config_window_edge_distance + self.config_window_edge_distance + self.line_edit_height)

    def handle_save_button_click(self):
        try:
            self.begin_line = int(self.begin_line_edit_text.text())
        except ValueError:
            self.begin_line = 2
            self.update_ui_signals.log_signal.emit("ERROR", "Begin line may illegal.")
        else:
            self.update_ui_signals.log_signal.emit("CONFIG", f"Update begin line = {self.begin_line_edit_text.text()} success.")
        if self.output_path_edit_text.text() != "":
            self.output_path = self.output_path_edit_text.text()
            self.update_ui_signals.log_signal.emit("CONFIG", f"Update output path = {self.output_path_edit_text.text()} success.")
        else:
            self.update_ui_signals.log_signal.emit("ERROR", "Path is null.")
        self.adjust_config_signals.adjust_output_path_signal.emit(self.output_path)
        self.adjust_config_signals.adjust_begin_line_signal.emit(self.begin_line)
        self.close()

    def handle_cancel_button_click(self):
        self.update_ui_signals.log_signal.emit("CONFIG", "Cancel.")
        self.close()


class VideoCrawler(QObject):
    app: QApplication
    windows: QWidget
    crawl_button: QPushButton
    import_button: QPushButton
    config_button: QPushButton
    progress_bar: QProgressBar
    log_box: QListWidget
    config_window: ConfigWindow

    window_width: int = 500
    window_height: int = 314
    button_width: int = 80
    button_height: int = 20
    edge_distance: int = 20

    ms_token = "g8vXKy2fjhjd7xrrPcCU7Wfop7isL5KAuyjofBp061Mtaxm_fA5vZ_lAlj46mvE_NnR7x-m-022QnOLdb6Em6HbYv1qm-ek2LXKHh6aTtnLpk_Ke8h7MUTkvWAUZX0cQ1JhrowY="

    file: str = ""
    task: dict = {}
    task_list: []
    progress: int = 0
    output_path: str = "/Users/rockey220224/Desktop/"
    file_name = ""
    begin_line: int = 2

    update_ui_signals: UpdateUISignals
    adjust_config_signals: AdjustConfigSignals

    feishu: Feishu

    def __init__(self):
        super().__init__()
        self.update_ui_signals = UpdateUISignals()
        self.adjust_config_signals = AdjustConfigSignals()
        self.feishu = Feishu(update_ui_signals=self.update_ui_signals)
        self.update_ui_signals.progress_bar_update_signal.connect(self.handle_update_progress_bar_signal)
        self.update_ui_signals.log_signal.connect(self.handle_log_signal)
        self.adjust_config_signals.adjust_begin_line_signal.connect(self.handle_adjust_begin_line)
        self.adjust_config_signals.adjust_output_path_signal.connect(self.handle_adjust_output_path)

    def render(self):
        self.app = QApplication(sys.argv)
        self.windows = QWidget()
        self.windows.resize(self.window_width, self.window_height)
        self.windows.setWindowTitle("短视频播放量爬虫（0.0）")
        self.windows.setFixedSize(self.windows.width(), self.windows.height())

        self.log_box = QListWidget(self.windows)
        self.log_box.move(self.edge_distance, self.edge_distance)
        self.log_box.resize(self.window_width - 2 * self.edge_distance, self.window_height - self.button_height - 3 * self.edge_distance)
        self.log_box.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

        self.crawl_button = QPushButton(self.windows)
        self.crawl_button.setText("开始爬取")
        self.crawl_button.move(int(self.window_width / 4 * 1 - self.button_width / 2), int(self.window_height - self.edge_distance - self.button_height))
        self.crawl_button.clicked.connect(self.handle_crawl_button_click)

        self.import_button = QPushButton(self.windows)
        self.import_button.setText("导入表格")
        self.import_button.move(int(self.window_width / 4 * 2 - self.button_width / 2), int(self.window_height - self.edge_distance - self.button_height))
        self.import_button.clicked.connect(self.handle_import_button_click)

        self.config_button = QPushButton(self.windows)
        self.config_button.setText("配置爬虫")
        self.config_button.move(int(self.window_width / 4 * 3 - self.button_width / 2), int(self.window_height - self.edge_distance - self.button_height))
        self.config_button.clicked.connect(self.handle_config_button_click)

        self.progress_bar = QProgressBar(self.windows)
        self.progress_bar.resize(self.window_width - 2 * self.edge_distance, 15)
        self.progress_bar.move(self.edge_distance, 3)

        self.windows.show()
        sys.exit(self.app.exec())
        pass

    def handle_crawl_button_click(self):
        def run():
            self.file_name = time.strftime("%Y-%m-%d_%H:%M:%S.xlsx", time.localtime())
            if self.task == {}:
                self.update_ui_signals.log_signal.emit("ERROR", "None file imported.")
            else:
                self.update_ui_signals.log_signal.emit("BEGIN", "Crawl start")
                wb = openpyxl.Workbook()
                if not self.check_output_file(wb):
                    return
                self.write_res_header(wb=wb)
                for url in self.task.get("task_list"):
                    url_type = self.get_url_type(url=url)
                    if url_type == "tiktok_without_video_id":
                        res = self.tiktok_without_video_id_proc(url)
                    elif url_type == "tiktok_with_video_id":
                        res = self.tiktok_with_video_id_proc(url)
                    else:
                        res = {
                            "found": -2,
                            "video_id": "",
                            "share": 0,
                            "play": 0,
                            "digg": 0,
                            "comment": 0,
                        }
                    if res is None:
                        res = {
                            "video_id": "",
                            "comment": 0,
                            "found": -3,
                            "digg": 0,
                            "play": 0,
                            "share": 0,
                        }
                    self.write_res_line(wb=wb,
                                        row=self.progress + 2,
                                        url=url,
                                        video_id=res["video_id"],
                                        comment=res["comment"],
                                        like=res["digg"],
                                        play=res["play"],
                                        share=res["share"],
                                        status=res["found"])
                    self.update_ui_signals.progress_bar_update_signal.emit()
                    self.progress = self.progress + 1
                self.update_ui_signals.log_signal.emit("COMPLETE", "Crawl complete.")

            pass

        crawl_thread: Thread = Thread(target=run)
        crawl_thread.start()
        pass

    def handle_import_button_click(self):
        res = QFileDialog.getOpenFileNames()[0]
        if len(res) != 0:
            self.file = res[0]
            self.log(log_type="FILE_IMPORT", log_text=f"Path:{self.file}.")
            try:
                self.task = self.read_xlsx()
            except InvalidFileException:
                self.log(log_type="ERROR", log_text=f"File '{self.file}' may not a xlsx file.")
            else:
                self.log(log_type="FILE_IMPORT", log_text="Xlsx file import success.")
        else:
            self.log(log_type="FILE_IMPORT", log_text="Cancel")
        pass

    def handle_config_button_click(self):
        self.config_window = ConfigWindow(begin_line=self.begin_line,
                                          output_path=self.output_path,
                                          adjust_config_signals=self.adjust_config_signals,
                                          update_ui_signals=self.update_ui_signals)
        self.config_window.show()

    def handle_log_signal(self, log_type, log_text):
        log_str: str = f"[{log_type}] {log_text}"

        def run(message: str):
            self.send_lark(message, log_type == "COMPLETE" and self.feishu.when_complete_at)

        send_thread: Thread = Thread(target=run, args=(log_str,))
        send_thread.start()
        self.log_box.addItem(log_str)
        self.log_box.verticalScrollBar().setValue(self.log_box.verticalScrollBar().maximum())
        pass

    def handle_update_progress_bar_signal(self):
        self.progress_bar.setValue(int(self.progress / (len(self.task["task_list"]) - 1) * 100))
        pass

    def handle_adjust_output_path(self, path: str):
        self.output_path = path
        pass

    def handle_adjust_begin_line(self, begin_line: int):
        self.begin_line = begin_line
        pass

    def log(self, log_type, log_text):
        log_str: str = f"[{log_type}] {log_text}"

        def run(message: str):
            self.feishu.send_lark(message)

        send_thread: Thread = Thread(target=run, args=(log_str,))
        send_thread.start()
        self.log_box.addItem(log_str)
        self.log_box.verticalScrollBar().setValue(self.log_box.verticalScrollBar().maximum())
        pass

    def read_xlsx(self) -> dict:
        wb: Workbook = openpyxl.open(filename=self.file)
        ws: Worksheet = wb.worksheets[0]
        self.task_list = []
        row: int = self.begin_line
        while ws.cell(row, 1).value is not None:
            self.task_list.append(ws.cell(row, 1).value)
            row = row + 1
        return {
            'task_count': row,
            'task_list': self.task_list
        }

    def get_abs_output_path(self) -> str:
        if self.output_path.endswith('/'):
            return self.output_path + self.file_name
        else:
            return self.output_path + '/' + self.file_name

    def write_res_line(self, wb: Workbook, row: int, url: str, video_id: str, status: int, like: int, comment: int, share: int, play: int):
        status_dict = {
            -1: "联系徐硕改代码",
            1: "Success",
            0: "Video lose efficacy",
            -2: "Bad URL or this website has no solution",
            -3: "Internet Error"
        }
        wb.worksheets[0].cell(row=row, column=1, value=url)
        wb.worksheets[0].cell(row=row, column=2, value=video_id)
        wb.worksheets[0].cell(row=row, column=3, value=status_dict[status])
        if status not in [-1, 0, -2, -3]:
            wb.worksheets[0].cell(row=row, column=4, value=like)
            wb.worksheets[0].cell(row=row, column=5, value=comment)
            wb.worksheets[0].cell(row=row, column=6, value=share)
            wb.worksheets[0].cell(row=row, column=7, value=play)
        wb.save(filename=self.get_abs_output_path())
        self.log(log_type="SAVE", log_text=f"Save id = {video_id} success.")

    def write_res_header(self, wb: Workbook, ):
        wb.worksheets[0].cell(row=1, column=1, value="URL")
        wb.worksheets[0].cell(row=1, column=2, value="VideoID")
        wb.worksheets[0].cell(row=1, column=3, value="Status")
        wb.worksheets[0].cell(row=1, column=4, value="LikesCount")
        wb.worksheets[0].cell(row=1, column=5, value="CommentsCount")
        wb.worksheets[0].cell(row=1, column=6, value="SharesCount")
        wb.worksheets[0].cell(row=1, column=7, value="PlaysCount")
        wb.save(filename=self.get_abs_output_path())
        self.log(log_type="SAVE", log_text="Save header success")

    def get_url_type(self, url: str):
        if url.startswith("https://vm.tiktok.com") or url.startswith("https://vt.tiktok.com"):
            return "tiktok_without_video_id"
        elif url.startswith("https://www.tiktok.com"):
            return "toktok_with_video_id"
        self.log(log_type="ERROR", log_text=f"Invalid URL '{url}'.")
        return ""

    def tiktok_without_video_id_proc(self, url: str):
        response = self.request_get(url=url, allow_redirects=False)
        if response is None:
            return
        soup = BeautifulSoup(response, "lxml")
        redirect_url = soup.find(name="a")["href"]
        video_id = self.get_tiktok_video_id(redirect_url)
        url = "https://www.tiktok.com/api/recommend/item_list/?aid=1988&app_language=es&app_name=tiktok_web&battery_info=0.44&browser_language=zh-CN&browser_name=Mozilla&browser_online=true&browser_platform=MacIntel&browser_version=5.0%20%28Macintosh%3B%20Intel%20Mac%20OS%20X%2010_15_7%29%20AppleWebKit%2F537.36%20%28KHTML%2C%20like%20Gecko%29%20Chrome%2F104.0.0.0%20Safari%2F537.36&channel=tiktok_web&cookie_enabled=true&count=30&device_id=7106463322559923714&device_platform=web_pc&focus_state=true&from_page=video&history_len=1&insertedItemID=" + video_id + "&is_fullscreen=true&is_page_visible=true&os=mac&priority_region=&referer=&region=DE&screen_height=1920&screen_width=1080&tz_name=Asia%2FShanghai&webcast_language=es&msToken=" + self.ms_token
        response = self.request_get(url=url, allow_redirects=False)
        if response is None:
            return
        rsp_json = json.loads(response)
        res = self.get_tiktok_data_from_api(except_id=video_id, rsp_json=rsp_json)
        if res["found"] == 1:
            self.update_ui_signals.log_signal.emit("SUCCESS", f"Get video:{video_id}; share: {res['share']}; play: {res['play']}; like: {res['digg']}; comment: {res['comment']}")
        elif res["found"] == 0:
            self.update_ui_signals.log_signal.emit("FAILED", "This video may lose efficacy")
        else:
            self.update_ui_signals.log_signal.emit("FAILED", "Api response may issue")
        return res

    def request_get(self, url: str, allow_redirects: bool):
        headers = {
            "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36"
        }
        try:
            response = requests.get(url, allow_redirects=allow_redirects, headers=headers, timeout=60)
        except MissingSchema:
            self.update_ui_signals.log_signal.emit("ERROR", f"Invalid URL '{url}', perhaps you meant 'https://{url}'.")
            return None
        except SSLError:
            self.update_ui_signals.log_signal.emit("ERROR", f"Max retries exceeded with URL '{url}'.")
            self.update_ui_signals.log_signal.emit("TIPS", "Let me have a rest.")
            self.update_ui_signals.log_signal.emit("TIPS", "Crawl will continue after 20s.")
            time.sleep(20)
            return None
        except requests.exceptions.ConnectionError:
            self.update_ui_signals.log_signal.emit("ERROR", "Internet error, maybe this error cause by VPN.")
            time.sleep(20)
            return None
        else:
            if response.status_code == 200 or response.status_code == 301:
                self.update_ui_signals.log_signal.emit("GET", f"{response.status_code}: {url}.")
                return response.content
            elif response.status_code == 404:
                return 404
            else:
                self.update_ui_signals.log_signal.emit("GET", f"{response.status_code}: {url}.")
                return None

    def get_tiktok_video_id(self, url: str) -> str:
        res_begin_index: int = url.rfind("/")
        res_end_index: int = url.find("?")
        res = url[res_begin_index + 1:res_end_index]
        for letter in res:
            if letter.isalpha():
                return ""
        return res

    def get_tiktok_data_from_api(self, except_id: str, rsp_json: dict) -> dict:
        try:
            item_list = rsp_json["itemList"]
            for item in item_list:
                if item["id"] == except_id:
                    share = item["stats"]["shareCount"]
                    play = item["stats"]["playCount"]
                    digg = item["stats"]["diggCount"]
                    comment = item["stats"]["commentCount"]
                    return {
                        "found": 1,
                        "video_id": except_id,
                        "share": share,
                        "play": play,
                        "digg": digg,
                        "comment": comment
                    }
            return {
                "found": 0,
                "video_id": except_id,
                "share": 0,
                "play": 0,
                "digg": 0,
                "comment": 0,
            }
        except KeyError:
            return {
                "found": -1,
                "video_id": except_id,
                "share": 0,
                "play": 0,
                "digg": 0,
                "comment": 0,
            }

    def tiktok_with_video_id_proc(self, url: str):
        video_id = self.get_tiktok_video_id(url)
        url_ = "https://www.tiktok.com/api/recommend/item_list/?aid=1988&app_language=es&app_name=tiktok_web&battery_info=0.44&browser_language=zh-CN&browser_name=Mozilla&browser_online=true&browser_platform=MacIntel&browser_version=5.0%20%28Macintosh%3B%20Intel%20Mac%20OS%20X%2010_15_7%29%20AppleWebKit%2F537.36%20%28KHTML%2C%20like%20Gecko%29%20Chrome%2F104.0.0.0%20Safari%2F537.36&channel=tiktok_web&cookie_enabled=true&count=30&device_id=7106463322559923714&device_platform=web_pc&focus_state=true&from_page=video&history_len=1&insertedItemID=" + video_id + "&is_fullscreen=true&is_page_visible=true&os=mac&priority_region=&referer=&region=DE&screen_height=1920&screen_width=1080&tz_name=Asia%2FShanghai&webcast_language=es&msToken=" + ms_token
        response = self.request_get(url=url_, allow_redirects=False)
        if response is None:
            return
        rsp_json = json.loads(response)
        res = self.get_tiktok_data_from_api(except_id=video_id, rsp_json=rsp_json)
        if res["found"] == 1:
            self.update_ui_signals.log_signal.emit("SUCCESS", f"Get video:{video_id}; share: {res['share']}; play: {res['play']}; like: {res['digg']}; comment: {res['comment']}.")
        elif res["found"] == 0:
            self.update_ui_signals.log_signal.emit("FAILED", "This video may lose efficacy")
        else:
            self.update_ui_signals.log_signal.emit("FAILED", "Api response may issue")
        return res

    def check_output_file(self, wb: Workbook) -> bool:
        try:
            wb.save(filename=self.get_abs_output_path())
        except FileNotFoundError:
            self.update_ui_signals.log_signal.emit("ERROR", "Output file not found. Please check folder exits:")
            self.update_ui_signals.log_signal.emit("ERROR", self.output_path)
            return False
        return True


class UserByHashtag(QObject):
    window_width: int = 500
    window_height: int = 314
    button_width: int = 80
    button_height: int = 20
    edge_distance: int = 20
    edit_line_x_distance = 80
    adjust_distance = 3

    task_list = []
    cur_line: int = 2

    output_path = "/Users/rockey211224/Desktop"
    file_name = ""

    app: QApplication
    windows: QWidget
    crawl_button: QPushButton
    verify_button: QPushButton
    log_box: QListWidget
    output_path_edit_text: QLineEdit
    hashtag_edit_text: QLineEdit
    output_path_label: QLabel
    hashtag_label: QLabel

    def __init__(self):
        super().__init__()
        self.update_ui_signals = UpdateUISignals()
        self.feishu = Feishu(update_ui_signals=self.update_ui_signals)
        self.update_ui_signals.log_signal.connect(self.handle_log_signal)

    def render(self):
        self.app = QApplication(sys.argv)
        self.windows = QWidget()
        self.windows.resize(self.window_width, self.window_height)
        self.windows.setWindowTitle("Hashtag用户信息爬虫（0.0）")
        self.windows.setFixedSize(self.windows.width(), self.windows.height())

        self.log_box = QListWidget(self.windows)
        self.log_box.move(self.edge_distance, self.edge_distance)
        self.log_box.resize(self.window_width - 2 * self.edge_distance, self.window_height - self.button_height - 4 * self.edge_distance)
        self.log_box.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAlwaysOff)

        self.crawl_button = QPushButton(self.windows)
        self.crawl_button.setText("开始爬取")
        self.crawl_button.move(int(self.window_width / 8 * 7 - self.button_width / 2), int(3 * self.edge_distance) + self.window_height - self.button_height - 4 * self.edge_distance - self.adjust_distance)
        self.crawl_button.clicked.connect(self.handle_crawl_button_click)

        self.verify_button = QPushButton(self.windows)
        self.verify_button.setText("格式校验")
        self.verify_button.move(int(self.window_width / 8 * 7 - self.button_width / 2), int(1.5 * self.edge_distance) + self.window_height - self.button_height - 4 * self.edge_distance - self.adjust_distance)
        self.verify_button.clicked.connect(self.handle_verify_button_click)

        self.output_path_label = QLabel(self.windows)
        self.output_path_label.setText("输出路径:")
        self.output_path_label.move(self.edge_distance, int(1.5 * self.edge_distance) + self.window_height - self.button_height - 4 * self.edge_distance)

        self.hashtag_label = QLabel(self.windows)
        self.hashtag_label.setText("Hashtag :")
        self.hashtag_label.move(self.edge_distance, int(3 * self.edge_distance) + self.window_height - self.button_height - 4 * self.edge_distance)

        self.output_path_edit_text = QLineEdit(self.windows)
        self.output_path_edit_text.move(self.edit_line_x_distance, int(1.5 * self.edge_distance) + self.window_height - self.button_height - 4 * self.edge_distance - self.adjust_distance)
        self.output_path_edit_text.setText(self.output_path)

        self.hashtag_edit_text = QLineEdit(self.windows)
        self.hashtag_edit_text.move(self.edit_line_x_distance, int(3 * self.edge_distance) + self.window_height - self.button_height - 4 * self.edge_distance - self.adjust_distance)

        self.windows.show()
        sys.exit(self.app.exec())
        pass

    def handle_crawl_button_click(self):
        if len(self.task_list) == 0:
            self.log("ERROR", "No task to run, input you hashtags and click '格式校验' button.")
            return

        def run():
            self.file_name = time.strftime("%Y-%m-%d_%H:%M:%S.xlsx", time.localtime())
            self.update_ui_signals.log_signal.emit("BEGIN", "Crawl start")
            wb = openpyxl.Workbook()
            if not self.check_output_file(wb):
                return
            self.write_head(wb=wb)
            for hashtag in self.task_list:
                self.proc_hashtag(hashtag=hashtag, wb=wb)
            self.update_ui_signals.log_signal.emit("COMPLETE", "Crawl complete.")
        crawl_thread: Thread = Thread(target=run)
        crawl_thread.start()

    def handle_verify_button_click(self):
        hashtags = self.hashtag_edit_text.text()
        self.__set_task(hashtags)
        pass

    def log(self, log_type, log_text):
        log_str: str = f"[{log_type}] {log_text}"

        def run(message: str):
            self.feishu.send_lark(message)

        send_thread: Thread = Thread(target=run, args=(log_str,))
        send_thread.start()
        self.log_box.addItem(log_str)
        self.log_box.verticalScrollBar().setValue(self.log_box.verticalScrollBar().maximum())
        pass

    def handle_log_signal(self, log_type, log_text):
        log_str: str = f"[{log_type}] {log_text}"

        def run(message: str):
            self.feishu.send_lark(message, log_type == "COMPLETE" and self.feishu.when_complete_at)

        send_thread: Thread = Thread(target=run, args=(log_str,))
        send_thread.start()
        self.log_box.addItem(log_str)
        self.log_box.verticalScrollBar().setValue(self.log_box.verticalScrollBar().maximum())
        pass

    def __set_task(self, hashtags: str):
        if len(hashtags) == 0:
            self.log("ERROR", "You should input something to hashtag.")
        hashtags = hashtags.replace(" ", "")
        hashtags = hashtags.replace("\n", "")
        self.task_list = hashtags.split("#")
        if "" in self.task_list:
            self.task_list.remove("")
        if len(self.task_list) == 0:
            self.log("ERROR", "There is no valid task to run, check your input.")
        else:
            self.log("ANALYSIS", "Analysis your input success.")
            self.log("ANALYSIS", f"Result: {self.task_list}")
        pass

    def request_get(self, url: str, allow_redirects: bool):
        headers = {
            "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.0.0 Safari/537.36"
        }
        try:
            response = requests.get(url, allow_redirects=allow_redirects, headers=headers, timeout=60)
        except MissingSchema:
            self.update_ui_signals.log_signal.emit("ERROR", f"Invalid URL '{url}', perhaps you meant 'https://{url}'.")
            return None
        except SSLError:
            self.update_ui_signals.log_signal.emit("ERROR", f"Max retries exceeded with URL '{url}'.")
            self.update_ui_signals.log_signal.emit("TIPS", "Let me have a rest.")
            self.update_ui_signals.log_signal.emit("TIPS", "Crawl will continue after 20s.")
            time.sleep(20)
            return None
        except requests.exceptions.ConnectionError:
            self.update_ui_signals.log_signal.emit("ERROR", "Internet error, maybe this error cause by VPN.")
            time.sleep(20)
            return None
        else:
            if response.status_code == 200 or response.status_code == 301:
                self.update_ui_signals.log_signal.emit("GET", f"{response.status_code}: {url}.")
                return response.content
            elif response.status_code == 404:
                return 404
            else:
                self.update_ui_signals.log_signal.emit("GET", f"{response.status_code}: {url}.")
                return None

    @staticmethod
    def timestamp_format(timestamp: int) -> str:
        create_time = time.localtime(timestamp)
        return str(time.strftime("%Y-%m-%d %H:%M:%S", create_time))

    def proc_hashtag(self, hashtag: str, wb: Workbook):
        self.log("PROGRESS", f"Begin crawl {hashtag}")
        hashtag_info: dict = self.get_hashtag_info(hashtag=hashtag)
        cursor = 0
        has_more = True
        while has_more:
            url = f'https://www.tiktok.com/api/challenge/item_list/?aid=1988&app_language=zh-Hant-TW&app_name=tiktok_web&battery_info=1&browser_language=zh-CN&browser_name=Mozilla&browser_online=true&browser_platform=MacIntel&browser_version=5.0%20%28Macintosh%3B%20Intel%20Mac%20OS%20X%2010_15_7%29%20AppleWebKit%2F537.36%20%28KHTML%2C%20like%20Gecko%29%20Chrome%2F104.0.0.0%20Safari%2F537.36&challengeID={hashtag_info["id"]}&channel=tiktok_web&cookie_enabled=true&count=30&cursor={cursor}&device_id=7135373539919021574&device_platform=web_pc&focus_state=true&from_page=hashtag&history_len=2&is_fullscreen=true&is_page_visible=true&language=zh-Hant-TW&os=mac&priority_region=&referer=https%3A%2F%2Fwww.google.com.hk%2F&region=TW&root_referer=https%3A%2F%2Fwww.google.com.hk%2F&screen_height=1920&screen_width=1080&tz_name=Asia%2FShanghai&webcast_language=zh-Hant-TW'
            cursor = cursor + 30
            rsp = self.request_get(url, allow_redirects=False)
            if rsp is None:
                continue
            rsp_json = json.loads(rsp)
            has_more = rsp_json["hasMore"]
            self.log("CRAWL", f"Has next page:{has_more}")
            if rsp_json.get("itemList") is None:
                print(rsp_json)
            if rsp_json["itemList"] is not None:
                self.log("CRAWL", f"Get {len(rsp_json['itemList'])} videos to crawl.")
                for video in rsp_json["itemList"]:
                    data = {
                        "Status": ResultStatus.Success,
                        "HashtagName": hashtag_info["title"],
                        "HashtagVideoCount": hashtag_info["videoCount"],
                        "HashtagViewCount": hashtag_info["viewCount"],
                        "UserNickname": video["author"]["nickname"],
                        "UserSignature": video["author"]["signature"],
                        "UserHomePage": f'https://www.tiktok.com/@{video["author"]["uniqueId"]}',
                        "UserDiggCount": video["authorStats"]["diggCount"],
                        "UserFollowerCount": video["authorStats"]["followerCount"],
                        "UserHeartCount": video["authorStats"]["heartCount"],
                        "UserVideoCount": video["authorStats"]["videoCount"],
                        "VideoPage": f'https://www.tiktok.com/@{video["author"]["uniqueId"]}/video/{video["id"]}',
                        "VideoDiggCount": video["stats"]["diggCount"],
                        "VideoPlayCount": video["stats"]["playCount"],
                        "VideoShareCount": video["stats"]["shareCount"],
                        "VideoCommentCount": video["stats"]["commentCount"],
                        "VideoCreateTime": self.timestamp_format(video["createTime"]),
                        "VideoDescription": video["desc"],
                        "VideoURL": video["video"].get("downloadAddr"),
                    }
                    if "@" in data.get("UserSignature"):
                        self.log("CRAWL", "Found '@' in user signature.")
                        self.write_line(wb=wb, data=data)
                    else:
                        self.log("CRAWL", "No found '@' in user signature.")

    def get_hashtag_info(self, hashtag: str) -> dict:
        url = "https://www.tiktok.com/api/challenge/detail/?aid=1988&app_language=zh-Hant-TW&app_name=tiktok_web&battery_info=1&browser_language=zh-CN&browser_name=Mozilla&browser_online=true&browser_platform=MacIntel&browser_version=5.0%20%28Macintosh%3B%20Intel%20Mac%20OS%20X%2010_15_7%29%20AppleWebKit%2F537.36%20%28KHTML%2C%20like%20Gecko%29%20Chrome%2F104.0.0.0%20Safari%2F537.36&challengeName=" + hashtag
        response_content = self.request_get(url=url, allow_redirects=False)
        if response_content is None:
            self.log("ERROR", f"Get {hashtag} hashtag info failed, this may cause by network.")
            return {
                "id": "-1",
                "title": hashtag,
                "videoCount": 0,
                "viewCount": 0,
            }
        rsp = json.loads(response_content)
        if rsp["statusCode"] == 10205:
            self.log("ERROR", f"No such hashtag: {hashtag}")
            return {
                "id": "-2",
                "title": hashtag,
                "videoCount": 0,
                "viewCount": 0,
            }
        hashtag_info = {
            "id": rsp["challengeInfo"]["challenge"]["id"],
            "title": rsp["challengeInfo"]["challenge"]["title"],
            "videoCount": rsp["challengeInfo"]["challenge"]["stats"]["videoCount"],
            "viewCount": rsp["challengeInfo"]["challenge"]["stats"]["viewCount"],
        }
        self.log("CRAWL", f"Get hashtag {hashtag} info success")
        self.log("CRAWL", f"id: {hashtag_info['id']}, title: {hashtag_info['title']}, videoCount: {hashtag_info['videoCount']}, viewCount: {hashtag_info['viewCount']}")
        return hashtag_info

    def write_head(self, wb: Workbook):
        wb.worksheets[0].cell(row=1, column=1, value="Status")
        wb.worksheets[0].cell(row=1, column=2, value="HashtagName")
        wb.worksheets[0].cell(row=1, column=3, value="HashtagVideoCount")
        wb.worksheets[0].cell(row=1, column=4, value="HashtagViewCount")
        wb.worksheets[0].cell(row=1, column=5, value="UserNickname")
        wb.worksheets[0].cell(row=1, column=6, value="UserSignature")
        wb.worksheets[0].cell(row=1, column=7, value="UserHomePage")
        wb.worksheets[0].cell(row=1, column=8, value="UserDiggCount")
        wb.worksheets[0].cell(row=1, column=9, value="UserFollowerCount")
        wb.worksheets[0].cell(row=1, column=10, value="UserHeartCount")
        wb.worksheets[0].cell(row=1, column=11, value="UserVideoCount")
        wb.worksheets[0].cell(row=1, column=12, value="VideoPage")
        wb.worksheets[0].cell(row=1, column=13, value="VideoDiggCount")
        wb.worksheets[0].cell(row=1, column=14, value="VideoPlayCount")
        wb.worksheets[0].cell(row=1, column=15, value="VideoShareCount")
        wb.worksheets[0].cell(row=1, column=16, value="VideoCommentCount")
        wb.worksheets[0].cell(row=1, column=17, value="VideoCreateTime")
        wb.worksheets[0].cell(row=1, column=18, value="VideoDescription")
        wb.worksheets[0].cell(row=1, column=19, value="VideoURL")
        wb.save(filename=self.get_abs_output_path())
        self.log(log_type="SAVE", log_text="Save header success")
        pass

    def write_line(self, wb: Workbook, data: dict):
        wb.worksheets[0].cell(row=self.cur_line, column=1, value=data["Status"])
        if data["Status"] == ResultStatus.Success:
            wb.worksheets[0].cell(row=self.cur_line, column=2, value=data["HashtagName"])
            wb.worksheets[0].cell(row=self.cur_line, column=3, value=data["HashtagVideoCount"])
            wb.worksheets[0].cell(row=self.cur_line, column=4, value=data["HashtagViewCount"])
            wb.worksheets[0].cell(row=self.cur_line, column=5, value=data["UserNickname"])
            wb.worksheets[0].cell(row=self.cur_line, column=6, value=data["UserSignature"])
            wb.worksheets[0].cell(row=self.cur_line, column=7, value=data["UserHomePage"])
            wb.worksheets[0].cell(row=self.cur_line, column=8, value=data["UserDiggCount"])
            wb.worksheets[0].cell(row=self.cur_line, column=9, value=data["UserFollowerCount"])
            wb.worksheets[0].cell(row=self.cur_line, column=10, value=data["UserHeartCount"])
            wb.worksheets[0].cell(row=self.cur_line, column=11, value=data["UserVideoCount"])
            wb.worksheets[0].cell(row=self.cur_line, column=12, value=data["VideoPage"])
            wb.worksheets[0].cell(row=self.cur_line, column=13, value=data["VideoDiggCount"])
            wb.worksheets[0].cell(row=self.cur_line, column=14, value=data["VideoPlayCount"])
            wb.worksheets[0].cell(row=self.cur_line, column=15, value=data["VideoShareCount"])
            wb.worksheets[0].cell(row=self.cur_line, column=16, value=data["VideoCommentCount"])
            wb.worksheets[0].cell(row=self.cur_line, column=17, value=data["VideoCreateTime"])
            wb.worksheets[0].cell(row=self.cur_line, column=18, value=data["VideoDescription"])
            wb.worksheets[0].cell(row=self.cur_line, column=19, value=data["VideoURL"])
        wb.save(filename=self.get_abs_output_path())
        self.log(log_type="SAVE", log_text=f"Save data success:{data}")
        self.cur_line = self.cur_line + 1

    def check_output_file(self, wb: Workbook) -> bool:
        try:
            wb.save(filename=self.get_abs_output_path())
        except FileNotFoundError:
            self.update_ui_signals.log_signal.emit("ERROR", "Output file not found. Please check folder exits:")
            self.update_ui_signals.log_signal.emit("ERROR", self.output_path)
            return False
        return True

    def get_abs_output_path(self) -> str:
        if self.output_path.endswith('/'):
            return self.output_path + self.file_name
        else:
            return self.output_path + '/' + self.file_name
